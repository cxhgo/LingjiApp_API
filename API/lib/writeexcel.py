# _*_ coding:utf-8 _*_
import os,sys
sys.path.append(os.path.dirname(os.path.dirname(__file__)))# window环境适配的代码
#sys.path.append(os.path.abspath('..'))# linux环境适配的代码
import shutil
from API.config import setting
from openpyxl import load_workbook
from openpyxl.styles import Font,Alignment
from openpyxl.styles.colors import RED,GREEN,DARKYELLOW
import configparser as cparser
# --------- 读取config.ini配置文件 ---------------
cf = cparser.ConfigParser()
cf.read(setting.Test_Config,encoding='UTF-8')
name = cf.get("tester","name")

class WriteExcel():
    # -------------文件写入数据---------------
    def __init__(self,filename,sheetname):
        """
        初始化数据
        :param filename:xls文件名
        :param sheetname:表格名称
        :return: 无
        """
        self.filename = filename
        # 如果不存在文件，则将文件1覆盖到文件2
        if not os.path.exists(self.filename):
            shutil.copyfile(setting.Source_File,setting.Target_File)
        # 打开文件表格，获取文件表
        self.wb = load_workbook(self.filename)
        self.ws = self.wb[sheetname]


    def write_data(self,row_num,value):
        """
        将测试结果和测试人员写入表格
        :param row_num:对应行数
        :param value:测试结果值
        :return: 无
        """
        # 写入测试结果，定义颜色、显示
        font_Creen = Font(name='宋体',color= GREEN,bold=True)
        font_Red = Font(name='宋体',color=RED,bold=True)
        font_Yellow = Font(name='宋体',color=DARKYELLOW,bold=True)
        align = Alignment(horizontal='center',vertical='center')
        # 获取写入结果的所在行数
        L_num = "L"+str(row_num)
        M_num = "M"+str(row_num)
        # 在第row_n行第12列写入value
        if value =="PASS":
            self.ws.cell(row_num,12,value)
            self.ws[L_num].font = font_Creen
        if value =="FAIL":
            self.ws.cell(row_num,12,value)
            self.ws[L_num].font = font_Red
        # 在第row_num行第13列写入测试人员名字
        self.ws.cell(row_num,13,name)
        self.ws[L_num].alignment =align
        self.ws[M_num].alignment = align
        self.ws[M_num].font = font_Yellow
        self.wb.save(self.filename)
